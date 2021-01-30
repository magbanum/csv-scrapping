import openpyxl
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string


if __name__ == '__main__':

    # read specific columns of csv file using Pandas 
    df = pd.read_csv(input("Enter csv file name "), usecols = ['Attendance for:','eds']) #column_1: 'Attendance for:'   column_2: 'eds'
    names=[]
    for i in range(2, 132):
        #print(repr(df['eds'][i]))
        if len(df['eds'][i]) == len(" u'\N{check mark}'") : #" ✔":
            names.append(df['Attendance for:'][i]) # participants
    
    #For loop to print names of present participant
    for e in names[0:]: 
        print(e)
        #continue
    #no. of participants present
    n = len(names)

    #load spreadsheet containing Class list
    wb = load_workbook('Google_Attendance.xlsx') 
    sheet = wb['Attendance Sheet'] 
    
    for name in names:

        for box in range(2, sheet.max_row+1):
            # checks if the participant is present in the sheet
           
            if name.lower() == (sheet.cell(row=box, column=column_index_from_string('A')).value.lower()):
                 # marks present of that participant  
                sheet[f'B{box}'] = 'Present' 

    wb.save('Google_Attendance.xlsx')
    print(f'No. of participants : {n}')  # 1 participant adds up while taking attendance
    print('Attendance taken successfully!')
